"""匯出路由 (Exports router) —— Phase 10 Excel / PDF 匯出。

職責:
  GET /projects/{pid}/export.xlsx
      以 openpyxl 產生兩個工作表:
        "Tasks"   : task_id, name, duration, es, ef, ls, lf, float, critical,
                    %complete, budget, EV, AC (逐任務)。
        "Summary" : 專案 KPI (工期 / 任務數 / 要徑數) + EVM 指標
                    (BAC/PV/EV/AC/SV/CV/SPI/CPI/EAC)。
      StreamingResponse, media_type 為 xlsx (openpyxl) 正確型別。
  GET /projects/{pid}/export.pdf
      以 reportlab 產生: 排程任務表 + EVM KPI + 任何待處理風險預警事件。
      StreamingResponse, media_type application/pdf。

設計重點:
  * 重用 projects._get_project_or_404 / get_project (CPM 結果) 與 progress
    router 既有的基準線 / 進度 / EVM 助手, evm.compute_evm 為純函式。
  * 無基準線時「優雅降級」(graceful degradation): EVM 區段省略;
    %complete / AC 改由 task_progress 直接取得 (EV 則因無基準線預算而留空)。
  * 唯讀 (read-only): 不寫入、不派工; viewer 角色亦可存取 (不加 require_role)。
  * 租戶隔離: 由 _get_project_or_404 以 tenant_id 過濾 (sqlite); PostgreSQL
    另有 RLS。sync_event_log (風險事件) 無 RLS, 程式以 tenant_id 過濾。
"""

from __future__ import annotations

import functools
import io
import logging

import anyio
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.evm import compute_evm
from app.core.i18n import t
from app.core.risk_listener import RISK_PROVISION_SYNC_TYPE
from app.deps import TenantContext, get_db, verify_tenant
from app.models.orm import SyncEvent
from app.routers.progress import (
    _baseline_tasks_for_evm,
    _load_active_baseline,
    _load_progress,
    _progress_map_for_evm,
)
from app.routers.projects import _get_project_or_404, get_project
from app.schemas.evm import EvmResult
from app.schemas.schedule import ProjectOut

logger = logging.getLogger("cpm.routers.exports")

# openpyxl 產生的 .xlsx 正確 MIME 型別 (OOXML spreadsheet)。
XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

router = APIRouter(prefix="/projects", tags=["exports"])


# ---------------------------------------------------------------------------
# 內部工具
# ---------------------------------------------------------------------------
async def _gather_export_context(
    db: AsyncSession, project_id: str, ctx: TenantContext
) -> tuple[ProjectOut, EvmResult | None, list[dict]]:
    """彙整匯出所需資料: CPM 後的 ProjectOut、選配 EVM、待處理風險事件。

    回傳 (project_out, evm_or_none, pending_risk_events)。
      * project_out : 含 CPM 結果 (es/ef/ls/lf/float/critical) 的完整輸出。
      * evm_or_none : 有基準線時於 data_date=project_duration 計算的 EvmResult;
                      無基準線時為 None (優雅降級)。
      * pending_risk_events : 本租戶、本專案、status=PENDING 的 RISK_PROVISION
                      事件 payload 清單。
    """
    project_out = await get_project(project_id, ctx=ctx, db=db)

    evm: EvmResult | None = None
    # Batch 5 FEAT-3：與 GET /evm / dashboard Q3 一致的 active-else-latest 規則。
    baseline = await _load_active_baseline(db, project_id)
    if baseline is not None:
        baseline_tasks = _baseline_tasks_for_evm(baseline)
        progress = await _progress_map_for_evm(db, project_id)
        snapshot = baseline.snapshot or {}
        data_date = project_out.project_duration
        if not data_date:
            data_date = int(snapshot.get("project_duration", 0))
        evm = compute_evm(baseline_tasks, progress, int(data_date))

    pending = await _pending_risk_events(db, ctx.tenant_id, project_id)
    return project_out, evm, pending


async def _pending_risk_events(
    db: AsyncSession, tenant_id: str, project_id: str
) -> list[dict]:
    """取得本租戶、本專案待處理 (PENDING) 的風險預警事件 payload 清單。

    Batch 4 (PERF-3)：直接以 sync_event_log.project_id 欄位過濾 (走複合索引
    (tenant_id, sync_type, status))，不再撈出全部 payload JSON 於 Python 比對。
    舊列的 project_id 已由遷移 0003 / main.py sqlite ALTER 自 payload 回填。
    """
    result = await db.execute(
        select(SyncEvent.payload).where(
            SyncEvent.tenant_id == tenant_id,
            SyncEvent.sync_type == RISK_PROVISION_SYNC_TYPE,
            SyncEvent.status == "PENDING",
            SyncEvent.project_id == project_id,
        )
    )
    return [payload for (payload,) in result.all() if isinstance(payload, dict)]


def _fmt_num(value: float | None, digits: int = 2) -> str:
    """數值格式化 (None -> 'N/A')。"""
    if value is None:
        return "N/A"
    return f"{value:,.{digits}f}"


def _planned_dates(
    project_out: ProjectOut, es: int, ef: int
) -> tuple[str | None, str | None]:
    """Batch 3 (FEAT-2)：由 day_dates 取任務的計畫開工 / 完工 ISO 日期。

    與 ERP payload 同一慣例：開工 = 第 es 個工作天、完工 = 第 ef-1 個工作天
    (最後一個施作日)；零工期 (里程碑) 以開工日為完工日。day_dates 未提供
    (專案無 start_date) 或索引越界時回 (None, None) —— 呼叫端據此省略欄位。
    """
    dd = project_out.day_dates
    if not dd:
        return None, None
    start_idx = max(int(es or 0), 0)
    finish_idx = max(int(ef or 0) - 1, start_idx)
    if start_idx >= len(dd) or finish_idx >= len(dd):
        return None, None
    return dd[start_idx], dd[finish_idx]


# ---------------------------------------------------------------------------
# Excel (openpyxl)
# ---------------------------------------------------------------------------
def _build_xlsx(
    project_out: ProjectOut,
    evm: EvmResult | None,
    progress_rows: list,
) -> bytes:
    """以 openpyxl 組裝 Tasks + Summary 兩個工作表, 回傳 .xlsx bytes。

    %complete / EV / AC: 有 EVM 時以 per_task 拆解為準; 否則退回 task_progress
    (EV 因無基準線預算而留空)。budget 一律取自 task_progress (無則 0)。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    # 進度 (budget / %complete / AC) 與 EVM per_task 索引。
    prog_by_task = {r.task_id: r for r in progress_rows}
    evm_by_task = {b.task_id: b for b in (evm.per_task if evm else [])}

    # Batch 3 (FEAT-2)：專案有開工日期 (day_dates) 時加入 開工/完工 日期欄。
    has_dates = bool(project_out.day_dates)

    # Batch 5 (FEAT-1)：依 (wbs_code, task_id) 排序 (未分類 "" 排最前)，
    # 使匯出結果依 WBS 分組呈現 (不影響 project_out.tasks 本身順序)。
    sorted_tasks = sorted(
        project_out.tasks, key=lambda tk: (tk.wbs_code or "", tk.task_id)
    )

    wb = Workbook()

    # ---- Tasks 工作表 ----
    ws = wb.active
    ws.title = "Tasks"
    headers = [
        "wbs_code",
        "task_id",
        "name",
        "duration",
        "es",
        "ef",
        "ls",
        "lf",
        "float",
        "critical",
        "%complete",
        "budget",
        "EV",
        "AC",
    ]
    if has_dates:
        # 置於 ef 之後、ls 之前以外的尾端 —— 追加欄位不打亂既有欄序 (向下相容)。
        headers = headers + ["開工", "完工"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2C3E50")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    for task in sorted_tasks:
        tid = task.task_id
        evm_row = evm_by_task.get(tid)
        prog = prog_by_task.get(tid)

        if evm_row is not None:
            percent_complete = evm_row.percent_complete
            budget = evm_row.budget
            ev_val: float | None = evm_row.ev
            ac_val: float | None = evm_row.ac
        elif prog is not None:
            percent_complete = int(prog.percent_complete or 0)
            budget = float(prog.budget or 0.0)
            ev_val = None  # 無基準線預算 -> EV 不可信, 留空
            ac_val = float(prog.actual_cost or 0.0)
        else:
            percent_complete = 0
            budget = 0.0
            ev_val = None
            ac_val = None

        row = [
            task.wbs_code or "",
            tid,
            task.task_name or "",
            int(task.duration or 0),
            int(task.es or 0),
            int(task.ef or 0),
            int(task.ls or 0),
            int(task.lf or 0),
            int(task.float_time or 0),
            "YES" if task.is_critical else "",
            percent_complete,
            budget,
            ev_val if ev_val is not None else "",
            ac_val if ac_val is not None else "",
        ]
        if has_dates:
            # Batch 3 (FEAT-2)：開工 = day_dates[es]、完工 = day_dates[ef-1]。
            start_iso, finish_iso = _planned_dates(
                project_out, int(task.es or 0), int(task.ef or 0)
            )
            row += [start_iso or "", finish_iso or ""]
        ws.append(row)

    # 高亮要徑列 (紅底白字)。
    crit_fill = PatternFill("solid", fgColor="E74C3C")
    crit_font = Font(color="FFFFFF")
    for row_idx, task in enumerate(sorted_tasks, start=2):
        if task.is_critical:
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = crit_fill
                cell.font = crit_font

    # 自動欄寬 (粗略估計)。
    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 40)

    # ---- Summary 工作表 ----
    summary = wb.create_sheet(title="Summary")
    bold = Font(bold=True)

    summary.append(["Project KPIs", ""])
    summary.cell(row=1, column=1).font = bold
    summary.append(["project_id", project_out.project_id])
    summary.append(["project_name", project_out.project_name])
    summary.append(["region", project_out.region])
    summary.append(["task_count", len(project_out.tasks)])
    summary.append(["project_duration", int(project_out.project_duration or 0)])
    summary.append(
        ["critical_count", sum(1 for tk in project_out.tasks if tk.is_critical)]
    )
    summary.append(["has_baseline", "YES" if evm is not None else "NO"])
    summary.append(["", ""])

    if evm is not None:
        evm_header = summary.max_row + 1
        summary.cell(row=evm_header, column=1, value="EVM (Earned Value)").font = bold
        summary.append(["data_date", evm.data_date])
        summary.append(["BAC", evm.bac])
        summary.append(["PV", evm.pv])
        summary.append(["EV", evm.ev])
        summary.append(["AC", evm.ac])
        summary.append(["SV", evm.sv])
        summary.append(["CV", evm.cv])
        summary.append(["SPI", evm.spi if evm.spi is not None else "N/A"])
        summary.append(["CPI", evm.cpi if evm.cpi is not None else "N/A"])
        summary.append(["EAC", evm.eac if evm.eac is not None else "N/A"])
        summary.append(["VAC", evm.vac if evm.vac is not None else "N/A"])
        summary.append(["risk_flagged", "YES" if evm.risk_flagged else "NO"])

    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 32

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    buf.close()
    return data


# ---------------------------------------------------------------------------
# PDF (reportlab)
# ---------------------------------------------------------------------------
def _build_pdf(
    project_out: ProjectOut,
    evm: EvmResult | None,
    pending_risk_events: list[dict],
    region: str,
) -> bytes:
    """以 reportlab 組裝: 排程表 + EVM KPI + 待處理風險事件, 回傳 PDF bytes。"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    # 重用 reports 的 CJK 字型註冊 (best-effort)。
    from app.automation.reports import _register_cjk_font

    region = (region or "TW").upper()
    font = _register_cjk_font()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        title=t(region, "reportTitle"),
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ExpTitle", parent=styles["Title"], fontName=font, fontSize=17, leading=21,
        spaceAfter=6,
    )
    meta_style = ParagraphStyle(
        "ExpMeta", parent=styles["Normal"], fontName=font, fontSize=10, leading=14,
        textColor=colors.HexColor("#555555"),
    )
    section_style = ParagraphStyle(
        "ExpSection", parent=styles["Heading2"], fontName=font, fontSize=12, leading=15,
        spaceBefore=8, spaceAfter=4, textColor=colors.HexColor("#2c3e50"),
    )
    cell_style = ParagraphStyle(
        "ExpCell", parent=styles["Normal"], fontName=font, fontSize=8.5, leading=11,
    )

    story: list = []

    # ---- 標題 + 專案資訊 ----
    story.append(Paragraph(t(region, "reportTitle"), title_style))
    meta_bits = [
        f"{t(region, 'project')}: {project_out.project_name} ({project_out.project_id})",
        f"{t(region, 'region')}: {project_out.region}",
        f"{t(region, 'projectDuration')}: {int(project_out.project_duration or 0)} {t(region, 'days')}",
    ]
    story.append(Paragraph("&nbsp;&nbsp;|&nbsp;&nbsp;".join(meta_bits), meta_style))
    story.append(Spacer(1, 6 * mm))

    # ---- 排程任務表 ----
    # Batch 3 (FEAT-2)：專案有開工日期 (day_dates) 時加入 計畫開工/完工 日期欄。
    has_dates = bool(project_out.day_dates)
    # Batch 5 (FEAT-1)：依 (wbs_code, task_id) 排序 (未分類 "" 排最前)。
    sorted_tasks = sorted(
        project_out.tasks, key=lambda tk: (tk.wbs_code or "", tk.task_id)
    )
    header = [
        Paragraph(t(region, "wbsCode"), cell_style),
        Paragraph(t(region, "taskId"), cell_style),
        Paragraph(t(region, "taskName"), cell_style),
        Paragraph(t(region, "duration"), cell_style),
        Paragraph("ES", cell_style),
        Paragraph("EF", cell_style),
        Paragraph("LS", cell_style),
        Paragraph("LF", cell_style),
        Paragraph(t(region, "floatTime"), cell_style),
        Paragraph(t(region, "critical"), cell_style),
    ]
    if has_dates:
        header += [
            Paragraph(t(region, "plannedStart"), cell_style),
            Paragraph(t(region, "plannedFinish"), cell_style),
        ]
    rows: list[list] = [header]
    critical_row_indexes: list[int] = []
    for idx, task in enumerate(sorted_tasks, start=1):
        if task.is_critical:
            critical_row_indexes.append(idx)
        row = [
            Paragraph(str(task.wbs_code or "-"), cell_style),
            Paragraph(str(task.task_id), cell_style),
            Paragraph(str(task.task_name or ""), cell_style),
            Paragraph(str(int(task.duration or 0)), cell_style),
            Paragraph(str(int(task.es or 0)), cell_style),
            Paragraph(str(int(task.ef or 0)), cell_style),
            Paragraph(str(int(task.ls or 0)), cell_style),
            Paragraph(str(int(task.lf or 0)), cell_style),
            Paragraph(str(int(task.float_time or 0)), cell_style),
            Paragraph("🔥" if task.is_critical else "-", cell_style),
        ]
        if has_dates:
            start_iso, finish_iso = _planned_dates(
                project_out, int(task.es or 0), int(task.ef or 0)
            )
            row += [
                Paragraph(start_iso or "-", cell_style),
                Paragraph(finish_iso or "-", cell_style),
            ]
        rows.append(row)
    if has_dates:
        # 日期雙欄需要寬度 —— 壓縮名稱欄與數值欄，總寬維持在版心內。
        col_widths = [
            14 * mm, 16 * mm, 26 * mm, 12 * mm, 10 * mm, 10 * mm, 10 * mm, 10 * mm,
            12 * mm, 12 * mm, 22 * mm, 22 * mm,
        ]
    else:
        col_widths = [
            16 * mm, 18 * mm, 38 * mm, 16 * mm, 14 * mm, 14 * mm, 14 * mm, 14 * mm,
            18 * mm, 16 * mm,
        ]
    table = Table(rows, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), font),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # Batch 5：新增 wbs_code 欄 (index 0) 使 duration 起的數值欄後移至 index 3。
        ("ALIGN", (3, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#bdc3c7")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f7fa")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    for r in critical_row_indexes:
        style_cmds.append(("BACKGROUND", (0, r), (-1, r), colors.HexColor("#e74c3c")))
        style_cmds.append(("TEXTCOLOR", (0, r), (-1, r), colors.white))
    table.setStyle(TableStyle(style_cmds))
    story.append(table)
    story.append(Spacer(1, 6 * mm))

    # ---- EVM KPI (有基準線才呈現) ----
    if evm is not None:
        story.append(Paragraph(t(region, "spi") + " / " + t(region, "cpi"), section_style))
        evm_rows = [
            [Paragraph(t(region, "bac"), cell_style), Paragraph(_fmt_num(evm.bac), cell_style),
             Paragraph("PV", cell_style), Paragraph(_fmt_num(evm.pv), cell_style)],
            [Paragraph("EV", cell_style), Paragraph(_fmt_num(evm.ev), cell_style),
             Paragraph("AC", cell_style), Paragraph(_fmt_num(evm.ac), cell_style)],
            [Paragraph(t(region, "scheduleVariance"), cell_style), Paragraph(_fmt_num(evm.sv), cell_style),
             Paragraph(t(region, "costVariance"), cell_style), Paragraph(_fmt_num(evm.cv), cell_style)],
            [Paragraph(t(region, "spi"), cell_style), Paragraph(_fmt_num(evm.spi, 3), cell_style),
             Paragraph(t(region, "cpi"), cell_style), Paragraph(_fmt_num(evm.cpi, 3), cell_style)],
            [Paragraph(t(region, "eac"), cell_style), Paragraph(_fmt_num(evm.eac), cell_style),
             Paragraph("VAC", cell_style), Paragraph(_fmt_num(evm.vac), cell_style)],
        ]
        evm_table = Table(evm_rows, colWidths=[40 * mm, 47 * mm, 40 * mm, 47 * mm])
        evm_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), font),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#bdc3c7")),
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ecf0f1")),
                    ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#ecf0f1")),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(evm_table)
        story.append(Spacer(1, 6 * mm))

    # ---- 待處理風險預警事件 ----
    if pending_risk_events:
        story.append(Paragraph(t(region, "riskProvision"), section_style))
        risk_rows = [
            [Paragraph("#", cell_style), Paragraph("reason", cell_style),
             Paragraph("detail", cell_style)]
        ]
        for i, ev in enumerate(pending_risk_events, start=1):
            reason = str(ev.get("reason", "")) if isinstance(ev, dict) else ""
            detail = ev.get("detail", {}) if isinstance(ev, dict) else {}
            detail_txt = ", ".join(
                f"{k}={v}" for k, v in (detail.items() if isinstance(detail, dict) else [])
            )
            risk_rows.append(
                [
                    Paragraph(str(i), cell_style),
                    Paragraph(reason, cell_style),
                    Paragraph(detail_txt or "-", cell_style),
                ]
            )
        risk_table = Table(risk_rows, colWidths=[12 * mm, 50 * mm, 112 * mm], repeatRows=1)
        risk_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), font),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#c0392b")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#bdc3c7")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(risk_table)

    doc.build(story)
    data = buf.getvalue()
    buf.close()
    return data


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------
@router.get("/{project_id}/export.xlsx")
async def export_xlsx(
    project_id: str,
    ctx: TenantContext = Depends(verify_tenant),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """匯出專案排程 + EVM 為 Excel (.xlsx)。viewer 角色亦可 (唯讀)。"""
    await _get_project_or_404(db, project_id, ctx.tenant_id)
    project_out, evm, _pending = await _gather_export_context(db, project_id, ctx)
    progress_rows = await _load_progress(db, project_id)

    # openpyxl 組裝為 CPU/IO 密集的同步作業, 以工作執行緒執行避免阻塞 event loop。
    xlsx_bytes = await anyio.to_thread.run_sync(
        functools.partial(_build_xlsx, project_out, evm, progress_rows)
    )
    filename = f"export_{project_id}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{project_id}/export.pdf")
async def export_pdf(
    project_id: str,
    ctx: TenantContext = Depends(verify_tenant),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """匯出專案排程 + EVM + 待處理風險為 PDF。viewer 角色亦可 (唯讀)。"""
    project = await _get_project_or_404(db, project_id, ctx.tenant_id)
    project_out, evm, pending = await _gather_export_context(db, project_id, ctx)

    # reportlab 組裝為 CPU 密集的同步作業, 以工作執行緒執行避免阻塞 event loop。
    pdf_bytes = await anyio.to_thread.run_sync(
        functools.partial(_build_pdf, project_out, evm, pending, project.region)
    )
    filename = f"export_{project_id}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )
